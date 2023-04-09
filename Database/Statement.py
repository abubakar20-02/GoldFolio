import math
import os
import sqlite3
import subprocess
import uuid
from datetime import datetime, timedelta
import calendar

import openpyxl
import pandas as pd
from fpdf import FPDF

from Database import DBFunctions, SetUpFile
import win32com.client as win32


class Statement:
    def __init__(self):
        super().__init__()
        self.c = None
        self.conn = None
        self.Profile = None

    def setProfile(self, user):
        self.Profile = user

    def createExcelTemplate(self):
        try:
            column_names = ['Date_Added', 'Gold', 'BoughtFor', 'ProfitLoss']

            # Create an empty DataFrame with the column names
            df = pd.DataFrame(columns=column_names)

            # Save the DataFrame to an Excel file
            df.to_excel('Statement.xlsx', index=False)
        except:
            None

    def isFileFormatCorrect(self, FilePath):
        column_names = ['Date_Added', 'Gold', 'BoughtFor', 'ProfitLoss']
        # source = 'UserTemplate.xlsx'
        target = FilePath
        # shutil.copyfile(source,target)
        # os.system(target)

        sheet_name = 'Sheet1'

        path = target

        # Read the Excel file into a DataFrame
        df = pd.read_excel(path, sheet_name=sheet_name)

        columnnames = df.columns
        for col in columnnames:
            if not col in column_names:
                return False
        return True

    def deleteUser(self):
        self.__SetUpConnection()
        self.c.execute("DELETE FROM Statement WHERE User_ID=?", (self.Profile,))
        self.conn.commit()
        self.conn.close()

    # can't handle empty lines at the moment
    def ImportFromExcel(self, FilePath):
        # source = 'UserTemplate.xlsx'
        target = FilePath
        # shutil.copyfile(source,target)
        # os.system(target)

        sheet_name = 'Sheet1'

        path = target

        # Read the Excel file into a DataFrame
        df = pd.read_excel(path, sheet_name=sheet_name)

        # check if user id exists already then only add. use purity boughtfor gold to be sure its real number.
        # date_added to be a date.

        # Define the SQL query to insert the data into the table
        table_name = "Statement"

        # Loop through the rows in the DataFrame and insert them into the table
        for _, row in df.iterrows():
            values = tuple(row)

            # check if the data is correct. Gold, Purity and Bought for have to be numbers and not empty whereas
            # profit loss only needs to be a number, but it can be None.
            if not (self.CorrectNumberFormat(values[2]) and self.CorrectNumberFormat(
                    values[3])):
                continue

            # self, InvestmentId, Date_added, UserID, Gold, Purity, BoughtFor

            # checks if time is none or date is empty.
            if isinstance(values[0], datetime) and values[0] is pd.NaT or self.isEmpty(values[0]):
                self.addIntoTable(values[1], 0.0, values[2], values[3])
            else:
                if isinstance(values[0], datetime):
                    # if date is in the future then don't add it.
                    # if datetime.strptime(values[0].strftime("%Y-%m-%d"), '%Y-%m-%d').date() > datetime.now().date():
                    #     print("future")
                    #     continue
                    # convert date to Y-m-d format
                    self.addIntoTable(values[1], 0.0, values[2], values[3], Date=values[0].strftime("%Y-%m-%d"))

    def isEmpty(self, value):
        # value is a number and it is not none.
        if isinstance(value, (float, int)) and (math.isnan(value)):
            return True
        else:
            return False

    def CorrectNumberFormat(self, value):
        # value is a number and it is not none.
        if isinstance(value, (float, int)) and not (math.isnan(value)):
            return True
        else:
            return False

    # __ makes the method private
    def __SetUpConnection(self):
        self.conn = sqlite3.connect(SetUpFile.DBName)
        self.c = self.conn.cursor()

    def createTable(self):
        self.__SetUpConnection()
        self.c.execute('''
              CREATE TABLE IF NOT EXISTS Statement
              ([Investment_ID] VARCHAR PRIMARY KEY, [Date_added] DATE DEFAULT CURRENT_DATE ,[User_ID] VARCHAR,[Gold] REAL ,[Purity] REAL, [BoughtFor] REAL, [ProfitLoss] REAL,[Value_Change] REAL,
              FOREIGN KEY(User_ID) REFERENCES User(User_ID))
              ''')
        self.conn.commit()
        self.conn.close()

    def addIntoTable(self, Gold, Purity, BoughtFor, ProfitLoss, Transaction_ID=None, Date=None):
        """Takes in the investmentID , User ID, Gold in grams, Purity and the total price bought for"""
        Value_Change = BoughtFor * (ProfitLoss / 100)
        self.__SetUpConnection()
        if Date is None:
            Date = datetime.now().date()
        else:
            if datetime.strptime(Date, '%Y-%m-%d').date() > datetime.now().date():
                print("date is in future")
                return
        Error = True
        # loop until there is no error.
        while Error:
            Error = False
            if Transaction_ID is None:
                Transaction_ID = str(uuid.uuid4())
            try:
                self.c.execute('''
                      INSERT INTO Statement (Investment_ID,Date_Added, User_ID , Gold, Purity, BoughtFor, ProfitLoss,Value_Change)

                            VALUES
                            (?,?,?,?,?,?,?,?)
                      ''', (Transaction_ID, Date, self.Profile, Gold, Purity, BoughtFor, ProfitLoss, Value_Change))
            except sqlite3.Error as error:
                print(error)
                Error = True
        self.conn.commit()
        self.conn.close()

    def deleteTable(self):
        self.__SetUpConnection()
        # Disable foreign key constraints
        self.c.execute("PRAGMA foreign_keys = OFF")
        try:
            self.c.execute("DROP TABLE Statement")
            self.conn.commit()
        except sqlite3.Error as error:
            print(error)
        finally:
            # Enable foreign key constraints
            self.c.execute("PRAGMA foreign_keys = ON")
            self.conn.close()

    def insertIntoTable(self, InvestmentId, Date_added, UserID, Gold, Purity, BoughtFor):
        """Takes in the investmentID , User ID, Gold in grams, Purity and the total price bought for"""
        self.__SetUpConnection()
        try:
            self.c.execute('''
                  INSERT INTO Investment (Investment_ID,Date_added, User_ID , Gold, Purity, BoughtFor, ProfitLoss,Value_Change)

                        VALUES
                        (?,?,?,?,?,?,?)
                  ''', (InvestmentId, Date_added, UserID, Gold, Purity, BoughtFor, 0.00, 0.00))
            self.conn.commit()
        except sqlite3.Error as error:
            print(error)
        finally:
            self.conn.close()

    def showStatement(self):
        self.__SetUpConnection()
        self.c.execute('''
                    SELECT * FROM Statement WHERE User_ID = ?
                  ''', (self.Profile,))
        self.conn.commit()
        df = pd.DataFrame(self.c.fetchall(),
                          columns=['Investment_ID', 'User_ID', 'Gold', 'Purity', 'BoughtFor', 'ProfitLoss'])
        print(df)
        self.conn.close()

    def getData(self, User_ID, *Investment_ID):
        """Takes in user id to show statement and it can also take investment id to get record for the investment id"""
        self.__SetUpConnection()
        Data = None
        self.c.execute("SELECT COUNT(*) FROM Statement WHERE User_ID = ?", (User_ID,))
        Count = self.c.fetchone()[0]
        print(Count)
        if Count > 0:
            if not Investment_ID:
                self.c.execute("SELECT * FROM Statement WHERE User_ID = ? LIMIT 1 OFFSET ?",
                               (User_ID, Count - 1,))
            else:
                self.c.execute(
                    "SELECT * FROM Statement WHERE Investment_ID = ? AND User_ID = ? LIMIT 1 OFFSET ?",
                    (Investment_ID, User_ID, Count - 1,))
            Data = self.c.fetchone()
            if not Investment_ID:
                # needs id to delete, its fine as id is primary key and unique.
                self.c.execute(
                    "DELETE FROM Statement WHERE Investment_ID IN (SELECT Investment_ID FROM Statement WHERE User_ID = ? LIMIT 1 OFFSET ?)",
                    (User_ID, Count - 1,))
            else:
                self.c.execute(
                    "DELETE FROM Statement WHERE (SELECT * FROM Statement WHERE User_ID = ? LIMIT 1 OFFSET ?)",
                    (User_ID, Count - 1,))

        self.conn.commit()
        self.conn.close()
        return Data

    def convertToExcel(self, StartDate=None, EndDate=None, FilePath='output_file.xlsx'):
        self.__SetUpConnection()
        # Define the parameters for the query
        params = (self.Profile,)

        # Query the database and create a DataFrame
        sql = 'SELECT Date_added,Gold,BoughtFor,ProfitLoss,Value_Change FROM Statement WHERE User_ID= ?'
        if StartDate is not None:
            sql += f" AND Date_Added >= '{StartDate.strftime('%Y-%m-%d')}'"
        if EndDate is not None:
            sql += f" AND Date_Added <= '{EndDate.strftime('%Y-%m-%d')}'"
        df = pd.read_sql(sql, con=self.conn, params=params)

        df.to_excel(FilePath, index=False)
        self.conn.close()
        subprocess.Popen(['start', 'excel.exe', FilePath], shell=True)

    def addtoTable(self, Values):
        print(Values)
        self.__SetUpConnection()
        self.c.executemany('''
              INSERT INTO Statement (Investment_ID,Date_Added, User_ID , Gold, Purity, BoughtFor, ProfitLoss)

                    VALUES
                    (?,?,?,?,?,?,?)
              ''', Values)
        self.conn.commit()
        self.conn.close()

    def getTable(self, StartDate=None, EndDate=None):
        self.__SetUpConnection()
        try:
            sql = "SELECT * FROM Statement WHERE User_ID = ?"
            if StartDate is not None:
                sql += f" AND Date_Added >= '{StartDate.strftime('%Y-%m-%d')}'"
            if EndDate is not None:
                sql += f" AND Date_Added <= '{EndDate.strftime('%Y-%m-%d')}'"
            values = (self.Profile,)
            df = pd.read_sql(sql, self.conn, params=values)
            df = df.drop('User_ID', axis=1)
            df = df.drop('Purity', axis=1)
        except sqlite3.Error as error:
            print(error)
        finally:
            self.conn.close()
            return df

    def getInvestmentCount(self, StartDate=None, EndDate=None):
        self.__SetUpConnection()
        sql = "SELECT COUNT(User_ID) From Statement WHERE User_ID=?"
        if StartDate:
            # idk why I have to do this.
            sql += f" AND Date_Added >= '{StartDate}'"

        if EndDate:
            sql += f" AND Date_Added <= '{EndDate}'"
        self.c.execute(sql,
                       (self.Profile,))
        value = self.c.fetchone()[0]
        self.conn.close()
        return value

    def getMinMaxDates(self):
        self.__SetUpConnection()
        self.c.execute("SELECT MIN(Date_Added),Max(Date_Added) FROM Statement")
        results = self.c.fetchone()
        self.conn.close()
        return results

    def Overall(self, Column, StartDate=None, EndDate=None):
        if StartDate is None and EndDate is None:
            StartDate, EndDate = self.getMinMaxDates()
            if StartDate is None and EndDate is None:
                return None
            StartDate = datetime.strptime(StartDate, '%Y-%m-%d').date()
            EndDate = datetime.strptime(EndDate, '%Y-%m-%d').date()

        self.__SetUpConnection()
        delta = abs(EndDate - StartDate)
        if delta >= timedelta(days=366):
            return self.Yearly(Column, Start=StartDate, End=EndDate)
            # generate yearly
            print("years")
        elif delta <= timedelta(days=31):
            return self.Daily(Column, Start=StartDate, End=EndDate)
            # generate single
            print("single")
        else:
            return self.Monthly(Column, Start=StartDate, End=EndDate)
            # generate monthly
            print("months")
        self.conn.close()

    def generate_monthly_dict(self, start_date, end_date):
        result = {}
        current_date = start_date.replace(day=1)
        while current_date <= end_date:
            result[current_date.strftime('%Y-%m')] = 0
            current_date += timedelta(days=32)
            current_date = current_date.replace(day=1)
        print(result)
        return result

    def generate_yearly_dict(self, start_date, end_date):
        result = {}
        for year in range(start_date.year, end_date.year + 1):
            result[year] = 0
        print(result)
        return result

    def generate_daily_dict(self, start_date, end_date):
        result = {}
        date = start_date
        while date <= end_date:
            result[date] = 0
            date += timedelta(days=1)
        return result

    def getPositiveTradeSum(self, StartDate=None, EndDate=None):
        self.__SetUpConnection()
        sql = "SELECT SUM(Value_Change) FROM Statement WHERE ProfitLoss>0 AND User_ID=?"
        if StartDate:
            sql += f" AND Date_added >= '{StartDate}'"
        if EndDate:
            sql += f" AND Date_added <= '{EndDate}'"
        self.c.execute(sql, (self.Profile,))
        value = self.c.fetchone()[0]
        if value is None:
            value = 0
        self.conn.close()
        return value

    def getNegativeTradeSum(self, StartDate=None, EndDate=None):
        self.__SetUpConnection()
        sql = "SELECT SUM(Value_Change) FROM Statement WHERE ProfitLoss<0 AND User_ID=?"
        if StartDate:
            sql += f" AND Date_added >= '{StartDate}'"
        if EndDate:
            sql += f" AND Date_added <= '{EndDate}'"
        self.c.execute(sql, (self.Profile,))
        value = self.c.fetchone()[0]
        if value is None:
            value = 0
        value = -value
        self.conn.close()
        return value

    def getProfitLossData(self, year, month):
        days_in_month = calendar.monthrange(year, month)[1]
        StartDate = datetime(year, month, 1).date()
        EndDate = datetime(year, month, days_in_month).date()
        a = ("Profit", self.getPositiveTradeSum(StartDate=StartDate, EndDate=EndDate))
        b = ("Loss", self.getNegativeTradeSum(StartDate=StartDate, EndDate=EndDate))
        data = dict([a, b])
        return data

    def Yearly(self, ColumnName, Start=None, End=None):
        self.__SetUpConnection()
        sql = (
            "SELECT strftime('%Y', Date_Added) AS year, SUM({0}) AS total_value FROM Statement WHERE User_ID=?").format(
            ColumnName)
        if Start:
            # idk why I have to do this.
            sql += f" AND Date_Added >= '{Start.strftime('%Y-%m-%d')}'"

        if End:
            sql += f" AND Date_Added <= '{End.strftime('%Y-%m-%d')}'"

        sql += f"GROUP BY year"

        if Start and End:
            yearlydict = self.generate_yearly_dict(Start, End)

        print(sql)

        self.c.execute(sql, (self.Profile,))

        total = 0
        for rows in self.c.fetchall():
            year, sum = rows
            total = total + sum
            print(type(year))
            yearlydict[datetime.strptime(year, '%Y').year] = total

        print("----")
        print(yearlydict)
        print("----")

        self.conn.close()
        return yearlydict

    def Monthly(self, ColumnName, Start=None, End=None):
        self.__SetUpConnection()
        # might need to include trade cost too
        sql = (
            "SELECT strftime('%Y-%m', Date_Added) AS month, SUM({0}) AS total_value FROM Statement WHERE User_ID=?").format(
            ColumnName)
        if Start:
            # idk why I have to do this.
            sql += f" AND Date_Added >= '{Start.strftime('%Y-%m-%d')}'"

        if End:
            sql += f" AND Date_Added <= '{End.strftime('%Y-%m-%d')}'"

        sql += f"GROUP BY month"

        print(sql)

        if Start and End:
            monthlydict = self.generate_monthly_dict(Start, End)

        print(sql)

        self.c.execute(sql, (self.Profile,))
        total = 0
        for rows in self.c.fetchall():
            print("hmm")
            print(rows)
            month, sum = rows
            total = total + sum
            monthlydict[month] = total

        print("----")
        print(monthlydict)
        print("----")

        self.conn.close()
        return monthlydict

    def Daily(self, ColumnName, Start=None, End=None):
        sql = "SELECT SUM({0}) FROM Statement WHERE Date_Added = ? AND User_ID=?".format(ColumnName)

        sql1 = "SELECT DISTINCT Date_Added FROM Statement WHERE User_ID = ?"

        if Start:
            # idk why I have to do this.
            StartDate = Start - timedelta(days=1)
            sql1 += f" AND Date_Added >= '{StartDate}'"

        if End:
            sql1 += f" AND Date_Added <= '{End}'"

        sql1 += f" ORDER BY Date_Added ASC"

        # execute SQL query to get all dates
        self.c.execute(sql1, (self.Profile,))

        dictionary = self.generate_daily_dict(Start, End)
        dates = []
        date = Start
        while date <= End:
            dates.append(date)
            date += timedelta(days=1)

        # loop through dates and print the sum of values for each date
        sum = 0
        for date in dates:
            # format_str = '%Y-%m-%d'
            # date = datetime.strptime(date, format_str)
            self.c.execute(sql, (date, self.Profile))
            sum_of_values = self.c.fetchone()[0]
            if sum_of_values is None:
                sum_of_values = 0
            sum += sum_of_values
            dictionary[date] = sum
            # if Mode is None:
            # dictionary[date] = sum_of_values
            # else:
            #     dictionary[date] = sum

        # close database connection
        self.conn.close()
        print("----")
        print(dictionary)
        print("----")
        return dictionary

    def getSUM(self, ColumnName, StartDate=None, EndDate=None):
        self.__SetUpConnection()
        sql = "SELECT SUM({0}) FROM Statement WHERE User_ID =?".format(ColumnName)
        if StartDate:
            sql += f" AND Date_Added >= '{StartDate}'"

        if EndDate:
            sql += f" AND Date_Added <= '{EndDate}'"
        self.c.execute(sql, (self.Profile,))
        Sum = self.c.fetchone()[0]
        self.conn.close()
        if Sum is None:
            Sum = 0
        return Sum

    def formatToWeek(self, number):
        ranges = []
        start = 1
        end = 7
        while end < number:
            ranges.append((start, end))
            start += 7
            end += 7
        ranges.append((start, number))
        return ranges

    def getDatesInWeekFormatForMonthValueChange(self, year, month):
        positive = {}
        negative = {}
        days_in_month = calendar.monthrange(year, month)[1]
        print(self.formatToWeek(days_in_month))
        for start, end in self.formatToWeek(days_in_month):
            start_date = datetime(year, month, start).date()
            end_date = datetime(year, month, end).date()
            positive[f"{start}-{end}"] = self.getPositiveTradeSum(StartDate=start_date, EndDate=end_date)
            negative[f"{start}-{end}"] = self.getNegativeTradeSum(StartDate=start_date, EndDate=end_date)

        return positive, negative

    def getDatesInWeekFormatForMonth(self, Column, year, month):
        dict1 = {}
        days_in_month = calendar.monthrange(year, month)[1]
        print(self.formatToWeek(days_in_month))
        for start, end in self.formatToWeek(days_in_month):
            start_date = datetime(year, month, start).date()
            end_date = datetime(year, month, end).date()
            dict1[f"{start}-{end}"] = self.getSUM(Column, StartDate=start_date, EndDate=end_date)

        return dict1

    def getSum(self, Column, StartDate=None, EndDate=None):
        sql = "SELECT SUM({0}) FROM Statement WHERE User_ID=?".format(Column)
        if StartDate:
            sql += f" AND Date_Added >= '{StartDate}'"
        if EndDate:
            sql += f" AND Date_Added <= '{EndDate}'"

        self.__SetUpConnection()
        self.c.execute(sql, (self.Profile,))
        sum = self.c.fetchone()[0]
        if sum is None:
            sum = 0
        self.conn.close()
        return sum

    def getAvgProfitLoss(self, StartDate=None, EndDate=None):
        Sum = self.getSum("Value_Change", StartDate=StartDate, EndDate=EndDate)
        Gold = self.getSum("Gold", StartDate=StartDate, EndDate=EndDate)
        if Sum == 0:
            return 0
        if Gold == 0:
            return 0
        return Sum / Gold

    def PDF(self, FilePath, StartDate=None, EndDate=None):
        self.__SetUpConnection()
        # Define the parameters for the query
        params = (self.Profile,)

        # Query the database and create a DataFrame
        sql = 'SELECT Date_added,Gold,BoughtFor,ProfitLoss,Value_Change FROM Statement WHERE User_ID= ?'
        if StartDate is not None:
            sql += f" AND Date_Added >= '{StartDate.strftime('%Y-%m-%d')}'"
        if EndDate is not None:
            sql += f" AND Date_Added <= '{EndDate.strftime('%Y-%m-%d')}'"
        df = pd.read_sql(sql, con=self.conn, params=params)

        # Create a PDF document using the fpdf library
        pdf = MyPDF()
        pdf.add_page()
        pdf.alias_nb_pages()

        # Set the font and size of the text in the PDF document
        pdf.set_font("Arial", size=12)

        # Calculate the maximum width of the data in each column
        column_width = pdf.w / len(df.columns)

        # Create a list of equal column widths that fill the width of the page
        column_widths = [column_width] * len(df.columns)

        # Set the left margin to 0 to stretch the table to take the entire width of the page
        left_margin = 0

        # Add the DataFrame to the PDF document as a table
        pdf.set_x(left_margin)
        for i, column in enumerate(df.columns):
            pdf.cell(column_widths[i], 10, str(column), border=1)
        for index, row in df.iterrows():
            pdf.ln()
            pdf.set_x(left_margin)
            for i, column in enumerate(df.columns):
                pdf.cell(column_widths[i], 10, str(row[column]), border=1)

        # Save the PDF document to a file
        pdf.output(FilePath)
        self.conn.close()


class MyPDF(FPDF):
    def footer(self):
        # Add a footer to the bottom center of each page
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')
